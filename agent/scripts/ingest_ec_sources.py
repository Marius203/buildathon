"""Ingest the 6 EC source documents into kb chunks for embedding.

Reads from agent/kb/:
  - Examples of Questions with Answers.docx  → one chunk per Q&A pair
  - Festival Rules and Regulations.docx      → ~700-char chunks
  - How to purchase tickets.docx             → one chunk per step / section
  - Daily recommendations.docx               → one chunk per detected SECTION
  - Important Links.docx                     → one chunk with all URLs
  - 2025 Line-up schedule & descriptions.xlsx → one chunk per artist

Writes:  data/mock/kb_chunks.ec_sources.json
Run:     python -m scripts.ingest_ec_sources

Idempotent: fully regenerates output each run. Chunk ids are deterministic
so a rerun-then-embed_kb produces upserts, not duplicates.

The BRIEF pdf and any other files in kb/ are ignored — only the 6 sources above
are read by name.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

KB_DIR = Path("kb")
OUT_PATH = Path("data/mock/kb_chunks.ec_sources.json")

# ----- helpers --------------------------------------------------------------

RO_DIACRITICS = set("ăâîșțĂÂÎȘȚşţŞŢ")
RO_WORDS = {
    "este", "pentru", "cu", "din", "care", "să", "și", "nu", "la", "pe",
    "îți", "pot", "ce", "mai", "fără", "după", "două", "puțin", "lângă",
}


def detect_lang(text: str) -> str:
    """Naive RO/EN detection. Defaults to 'en'."""
    if set(text) & RO_DIACRITICS:
        return "ro"
    words = set(re.findall(r"[a-zăâîșț]+", text.lower())[:80])
    if len(words & RO_WORDS) >= 3:
        return "ro"
    return "en"


def slugify(s: str, max_len: int = 40) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s.lower()).strip("_")
    return s[:max_len] or "item"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _read_docx_paras(path: Path) -> list[str]:
    doc = Document(path)
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]


# ----- parsers --------------------------------------------------------------

def parse_faq() -> list[dict]:
    """Q&A pairs grouped by section header. Each pair becomes one chunk."""
    path = KB_DIR / "Examples of Questions with Answers.docx"
    if not path.exists():
        print(f"[warn] FAQ missing: {path}")
        return []

    paras = _read_docx_paras(path)

    # Trim to content between markers if present.
    try:
        start = paras.index("***Question examples start***") + 1
    except ValueError:
        start = 0
    try:
        end = paras.index("***Question examples end***")
    except ValueError:
        end = len(paras)
    paras = paras[start:end]

    chunks: list[dict] = []
    section = "general"
    q_idx = 0
    pending_q: str | None = None
    pending_a: list[str] = []
    ts = now_iso()

    def flush() -> None:
        nonlocal pending_q, pending_a, q_idx
        if pending_q is not None and pending_a:
            q_idx += 1
            answer = " ".join(pending_a).strip()
            text = f"Q: {pending_q}\nA: {answer}"
            chunks.append({
                "id": f"ec_faq_{slugify(section)}_q{q_idx}",
                "text": text,
                "lang": detect_lang(text),
                "topic": "faq",
                "section": slugify(section),
                "section_title": section,
                "source": "ec_faq",
                "created_at": ts,
            })
        pending_q = None
        pending_a = []

    for p in paras:
        if p.startswith("-"):
            flush()
            pending_q = p.lstrip("- ").strip()
        elif pending_q is not None:
            pending_a.append(p)
        else:
            # Section header: short, not a question, not following a question.
            if len(p) <= 60:
                section = p
                q_idx = 0
    flush()
    return chunks


def parse_rules() -> list[dict]:
    """Group consecutive paragraphs until ~700 chars, then flush."""
    path = KB_DIR / "Festival Rules and Regulations.docx"
    if not path.exists():
        print(f"[warn] rules missing: {path}")
        return []

    paras = _read_docx_paras(path)

    chunks: list[dict] = []
    buffer: list[str] = []
    buf_chars = 0
    idx = 0
    ts = now_iso()
    TARGET_CHARS = 700

    def flush() -> None:
        nonlocal buffer, buf_chars, idx
        if not buffer:
            return
        idx += 1
        text = "\n".join(buffer)
        chunks.append({
            "id": f"ec_rules_{idx:03d}",
            "text": text,
            "lang": detect_lang(text),
            "topic": "rules",
            "section": str(idx),
            "section_title": f"Rules chunk {idx}",
            "source": "ec_rules",
            "created_at": ts,
        })
        buffer = []
        buf_chars = 0

    for p in paras:
        buffer.append(p)
        buf_chars += len(p)
        if buf_chars >= TARGET_CHARS:
            flush()
    flush()
    return chunks


def parse_tickets() -> list[dict]:
    """Split by numbered step headers and the trailing labelled sections."""
    path = KB_DIR / "How to purchase tickets.docx"
    if not path.exists():
        print(f"[warn] tickets missing: {path}")
        return []

    paras = _read_docx_paras(path)

    step_re = re.compile(r"^\d+\.\s+")
    SECTION_MARKERS = {
        "How to Buy Electric Castle Tickets",
        "Step-by-Step Purchase Guide",
        "Important Rules",
        "Buy Only from Official Partners",
    }

    chunks: list[dict] = []
    section_title = "Overview"
    buffer: list[str] = []
    ts = now_iso()
    idx = 0

    def flush() -> None:
        nonlocal buffer, idx
        if not buffer:
            return
        idx += 1
        text = "\n".join(buffer)
        chunks.append({
            "id": f"ec_tickets_{idx:02d}_{slugify(section_title)}",
            "text": text,
            "lang": detect_lang(text),
            "topic": "tickets",
            "section": str(idx),
            "section_title": section_title,
            "source": "ec_tickets",
            "created_at": ts,
        })
        buffer = []

    for p in paras:
        if step_re.match(p) or p in SECTION_MARKERS:
            flush()
            section_title = p
            buffer = [p]
        else:
            buffer.append(p)
    flush()
    return chunks


def parse_daily_recs() -> list[dict]:
    """One chunk per known section header (days + topic categories)."""
    path = KB_DIR / "Daily recommendations.docx"
    if not path.exists():
        print(f"[warn] daily recs missing: {path}")
        return []

    paras = _read_docx_paras(path)

    KNOWN_HEADERS: dict[str, str] = {
        "RECOMMENDATIONS": "recommendations",
        "WEDNESDAY": "recommendations",
        "THURSDAY": "recommendations",
        "FRIDAY": "recommendations",
        "SATURDAY": "recommendations",
        "SUNDAY": "recommendations",
        "SAFETY & ESSENTIAL INFO": "safety",
        "CASHLESS REFUND": "cashless",
        "ACTIVATING THE WRISTBANK": "cashless",
        "PIZZA": "food",
        "BURGERS": "food",
        "MEAT": "food",
        "TEX-MEX": "food",
        "VEGAN": "food",
        "ETHNIC FOOD": "food",
        "FINE DINING": "food",
        "COMFORT FOOD": "food",
        "SWEETS": "food",
        "TRANSPORTATION": "transport",
        "SCHEDULE FOR EC TRAINS:": "transport",
        "BUS:": "transport",
        "MAP & LOCATIONS": "venue",
        "STAGES": "venue",
        "EXPERIENCES & ACTIVITIES": "experiences",
    }

    chunks: list[dict] = []
    section_title = "Intro"
    topic = "recommendations"
    buffer: list[str] = []
    ts = now_iso()

    def flush() -> None:
        nonlocal buffer
        if not buffer:
            return
        text = "\n".join(buffer)
        chunks.append({
            "id": f"ec_daily_{slugify(section_title)}",
            "text": text,
            "lang": detect_lang(text),
            "topic": topic,
            "section": slugify(section_title),
            "section_title": section_title,
            "source": "ec_daily_recs",
            "created_at": ts,
        })
        buffer = []

    for p in paras:
        if p in KNOWN_HEADERS:
            flush()
            section_title = p
            topic = KNOWN_HEADERS[p]
            buffer = []
        else:
            buffer.append(p)
    flush()
    return chunks


def parse_links() -> list[dict]:
    """All official links concatenated into one chunk."""
    path = KB_DIR / "Important Links.docx"
    if not path.exists():
        print(f"[warn] links missing: {path}")
        return []

    paras = _read_docx_paras(path)
    text = "\n".join(paras)
    ts = now_iso()
    return [{
        "id": "ec_links_all",
        "text": text,
        "lang": "en",
        "topic": "links",
        "section": "all",
        "section_title": "Important Links",
        "source": "ec_links",
        "created_at": ts,
    }]


def parse_lineup() -> list[dict]:
    """One chunk per artist from the 'Final lineup table' sheet."""
    path = KB_DIR / "2025 Line-up schedule & descriptions.xlsx"
    if not path.exists():
        print(f"[warn] lineup missing: {path}")
        return []

    wb = load_workbook(path, data_only=True)
    sheet_name = "Final lineup table"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    chunks: list[dict] = []
    ts = now_iso()
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header in 'Final lineup table' is: Artist Name | Time and place | Description | Genre | ...
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        artist = str(row[0]).strip()
        schedule = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        description = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        genres = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if not artist:
            continue

        text_parts = [f"Artist: {artist}"]
        if genres and genres != "#N/A":
            text_parts.append(f"Genre: {genres}")
        if schedule:
            text_parts.append(f"Schedule: {schedule}")
        if description:
            text_parts.append(f"Description: {description}")
        text = "\n".join(text_parts)

        chunks.append({
            "id": f"ec_lineup_{slugify(artist, max_len=60)}",
            "text": text,
            "lang": detect_lang(description or text),
            "topic": "lineup",
            "section": slugify(artist, max_len=60),
            "section_title": artist,
            "source": "ec_lineup_2025",
            "created_at": ts,
        })

    return chunks


# ----- main -----------------------------------------------------------------

def main() -> None:
    parsers = [
        ("FAQ",        parse_faq),
        ("Rules",      parse_rules),
        ("Tickets",    parse_tickets),
        ("Daily recs", parse_daily_recs),
        ("Links",      parse_links),
        ("Lineup",     parse_lineup),
    ]

    all_chunks: list[dict] = []
    for label, fn in parsers:
        try:
            chunks = fn()
        except Exception as exc:
            print(f"[err] {label}: {exc}", file=sys.stderr)
            raise
        print(f"[ok] {label:12s} {len(chunks):4d} chunks")
        all_chunks.extend(chunks)

    # Dedup by id (last write wins). Slugified ids can collide for similar names.
    seen: dict[str, dict] = {}
    for c in all_chunks:
        seen[c["id"]] = c
    final = list(seen.values())

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(final, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\n[done] wrote {len(final)} chunks -> {OUT_PATH}")

    by_source: dict[str, int] = {}
    by_topic: dict[str, int] = {}
    by_lang: dict[str, int] = {}
    for c in final:
        by_source[c["source"]] = by_source.get(c["source"], 0) + 1
        by_topic[c["topic"]] = by_topic.get(c["topic"], 0) + 1
        by_lang[c["lang"]] = by_lang.get(c["lang"], 0) + 1

    print("\n[summary] chunks per source:")
    for k, v in sorted(by_source.items(), key=lambda x: -x[1]):
        print(f"        {k:20s} {v}")
    print("\n[summary] chunks per topic:")
    for k, v in sorted(by_topic.items(), key=lambda x: -x[1]):
        print(f"        {k:20s} {v}")
    print("\n[summary] chunks per lang:")
    for k, v in sorted(by_lang.items(), key=lambda x: -x[1]):
        print(f"        {k:20s} {v}")

    if final:
        sizes = sorted(len(c["text"]) for c in final)
        print(
            f"\n[summary] text length (chars): "
            f"min={sizes[0]} median~{sizes[len(sizes)//2]} max={sizes[-1]}"
        )


if __name__ == "__main__":
    main()
